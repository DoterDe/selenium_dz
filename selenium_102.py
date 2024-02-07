from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
chrome= webdriver.Chrome(keep_alive=True)

chrome.get('https://www.saucedemo.com/')

username=chrome.find_element(By.ID, "user-name")
password = chrome.find_element(By.NAME, "password")

button = chrome.find_element(By.XPATH, '//*[@id="login-button"]')

username.send_keys("standard_user")
password.send_keys("secret_sauce")
button.click()

all = chrome.find_elements(By.CLASS_NAME, 'inventory_item_description')
name_product = chrome.find_elements(By.CLASS_NAME, 'inventory_item_name')
name_desc = chrome.find_elements(By.CLASS_NAME, 'inventory_item_desc')
name_price= chrome.find_elements(By.CLASS_NAME, 'inventory_item_price')
wb = Workbook()
ws = wb.active
ws['A1']="№"
ws["B1"]="name"
ws["C1"]="desc"
ws["D1"]="price"
with open('secret.xlsx', 'w' , encoding='utf-8')as f:
    for  i in range(len(name_product)):
        ws[f'a{i+2}']=i+1
        ws[f'B{i+2}']=name_product[i].text
        ws[f'C{i+2}']=name_desc[i].text
        ws[f'D{i+2}']=name_price[i].text
time.sleep(10)
wb.save('secret.xlsx')




























